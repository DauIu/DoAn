import json
import keyword
from datetime import date, datetime, timedelta
from decimal import Decimal

from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from django.views import View
from django.db.models import Max, Sum, F
from django.db import transaction
from django.core.paginator import Paginator
from django.views.decorators.csrf import csrf_exempt
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.utils import timezone
from django.utils.timezone import now, localtime
from django.utils.dateparse import parse_date
from django.core.exceptions import ValidationError
from django.dispatch import receiver
from django.db.models.signals import post_save
from django.views.decorators.http import require_POST 
from django.shortcuts import render
from django.core.mail import send_mail
from django.contrib import messages
from django.contrib.auth.hashers import make_password
from .models import DanhGia
import random
import string
import openpyxl
# Forms and Models
from .forms import (
    KhachHangForm, RegistrationForm, LoginForm, LoaiForm, SanPhamForm, UserForm, VoucherForm,
    NhaCungCapForm, PhieuNhapForm, CT_PhieuNhapForm, KhoForm, TonKhoForm, 
    TinTucForm, SanPhamHuForm, DanhGiaForm
)
from .models import (
    CT_DonHang, DonHang, SanPham, CustomUser, Loai, SanPhamHu, Voucher, 
    NhaCungCap, PhieuNhap, CT_PhieuNhap, Kho, TonKho, TinTuc as TinTucModel,DanhGia, Quan, Phuong, ThanhPho
)
from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode
from django.utils.encoding import force_bytes
from django.contrib.sites.shortcuts import get_current_site
from django.template.loader import render_to_string
from django.core.mail import send_mail
from django.contrib.auth import get_user_model
from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_decode
from django.contrib.auth import get_user_model
User = get_user_model()
# Create your views here.

#----search----
def search_products(request):
    query = request.GET.get('q', '')  # Lấy từ khóa từ query string
    results = []
    if query:
        # Tìm kiếm sản phẩm có tên chứa từ khóa (không phân biệt hoa thường)
        results = SanPham.objects.filter(TenSP__icontains=query)
    return render(request, 'search_results.html', {'results': results, 'query': query})

#---------- Trang thông tin
def home(request):
    # Lấy sản phẩm 
    sanphams_nhapkhau = SanPham.objects.filter(MaLoai_id=1)[:6]
    sanphams_viet = SanPham.objects.filter(MaLoai_id=2)[:6]
    sanphams_giotraicay = SanPham.objects.filter(MaLoai_id=3)[:6]

    #Lấy tin tức mới nhất
    tin_moi = TinTucModel.objects.all().order_by('-NgayDang')[:3]

    return render(request, 'pages/User/Home.html', {
        'sanphams_nhapkhau': sanphams_nhapkhau, 
        'sanphams_viet': sanphams_viet, 
        'sanphams_giotraicay': sanphams_giotraicay,
        'tin_moi': tin_moi
    })

def GioiThieu(request):
    return render(request, 'pages/User/GioiThieu.html')
def LienHe(request):
    return render(request, 'pages/User/LienHe.html')
def TinTuc(request):
    tin_tuc_list = TinTucModel.objects.all().order_by('-NgayDang') 
    paginator = Paginator(tin_tuc_list, 3) 
    page_number = request.GET.get('page') 
    page_obj = paginator.get_page(page_number)

    return render(request, 'pages/User/TinTuc.html', {'dsTinTuc': page_obj })

def Vouchers(request):
    today = now().date()
    voucher_list = Voucher.objects.all().order_by('-NgayKetThuc') 
    paginator = Paginator(voucher_list, 9) 
    page_number = request.GET.get('page') 
    page_obj = paginator.get_page(page_number)

    return render(request, 'pages/User/Voucher.html', {'dsVoucher': page_obj, 'today': today})

#---------- Đang nhập - Đăng ký - Quên mật khẩu
class register(View):
    def get(self, request):
        form = RegistrationForm()
        city_lst = ThanhPho.objects.all()  # Lấy tất cả thành phố từ cơ sở dữ liệu
        return render(request, 'pages/User/DangKy.html', {'form': form, 'city_lst': city_lst})

    def post(self, request):
        city_id = request.POST.get('city')
        district_id = request.POST.get('district')
        ward_id = request.POST.get('ward')

        # Cập nhật form với city_id và district_id
        form = RegistrationForm(request.POST, city_id=city_id, district_id=district_id)
        city_lst = ThanhPho.objects.all()

        if form.is_valid():            
            try:
                # Kiểm tra nếu ID thành phố, quận, và phường hợp lệ
                city = ThanhPho.objects.get(MaTP=city_id)
                district = Quan.objects.get(MaQuan=district_id)
                ward = Phuong.objects.get(MaPhuong=ward_id)
                
                user = CustomUser.objects.create_user(
                    email=form.cleaned_data['email'],
                    username=form.cleaned_data['username'],
                    phone=form.cleaned_data['phone'],
                    full_name=form.cleaned_data['full_name'],
                    street=form.cleaned_data['street'],
                    city=city,
                    district=district,
                    ward=ward,
                    password=form.cleaned_data['password1']
                )
                user.save()
                return redirect('UserMember:DangNhap')
            except ThanhPho.DoesNotExist:
                form.add_error('city', 'Thành phố không hợp lệ.')
            except Quan.DoesNotExist:
                form.add_error('district', 'Quận/Huyện không hợp lệ.')
            except Phuong.DoesNotExist:
                form.add_error('ward', 'Phường/Xã không hợp lệ.')

        else:
            print("Form errors:", form.errors)
        return render(request, 'pages/User/DangKy.html', {'form': form, 'city_lst': city_lst})

class UserLogin(View):
    def get(self, request):
        # Hiển thị form đăng nhập
        formDangNhap = LoginForm()
        return render(request, 'pages/User/DangNhap.html', {'formDangNhap': formDangNhap})

    def post(self, request):
        # Lấy thông tin từ form
        username = request.POST.get('username')
        password = request.POST.get('password')

        # Xác thực người dùng bằng username
        user = authenticate(request, username=username, password=password)

        if user is not None:
            # Kiểm tra trạng thái hoạt động của tài khoản
            if user.is_active:
                print("\u2714\ufe0f Đăng nhập thành công")

                # Cập nhật ngày đăng nhập cuối
                user.ngaydangnhap = now().date()
                user.save()

                # Đăng nhập người dùng
                login(request, user)

                # Điều hướng dựa trên quyền hạn
                if user.is_admin:  # Admin quyền cao nhất
                    return redirect('UserMember:qlHome')
                elif user.is_manager:  # Quản lý
                    return redirect('UserMember:qlHome')
                elif user.is_staff:  # Nhân viên
                    return redirect('UserMember:qlHome')  
                elif user.is_user:  # Người dùng thường
                    return redirect('UserMember:home')
                else:
                    print("\u274c Quyền không hợp lệ")
                    return self.render_with_error(request, "Người dùng không có quyền truy cập hợp lệ.")
            else:
                # Tài khoản không hoạt động
                print("\u274c Tài khoản không hoạt động")
                return self.render_with_error(request, "Tài khoản của bạn chưa được kích hoạt.")
        else:
            # Thông tin đăng nhập không chính xác
            print("\u274c Đăng nhập thất bại")
            return self.render_with_error(request, "Tên đăng nhập hoặc mật khẩu không chính xác.")

    def render_with_error(self, request, error_message):
        # Render lại form với thông báo lỗi
        formDangNhap = LoginForm()
        return render(request, 'pages/User/DangNhap.html', {
            'formDangNhap': formDangNhap,
            'error_message': error_message
        })
         
def UserLogout(request):
    logout(request)
    return redirect('UserMember:home')

def password_reset(request):
    return render(request, 'pages/User/password_reset.html')

#----------Chọn vị trí-----------------


class GetQuan(View):
    def get(self, request, city_id):
        # Truy vấn các quận theo thành phố
        quans = Quan.objects.filter(thanh_pho_id=city_id)
        quan_list = [{'id': quan.MaQuan, 'name': quan.TenQuan} for quan in quans]
        return JsonResponse({'districts': quan_list})

class GetPhuong(View):
    def get(self, request, district_id):
        # Truy vấn các phường theo quận
        phuongs = Phuong.objects.filter(quan_id=district_id)
        phuong_list = [{'id': phuong.MaPhuong, 'name': phuong.TenPhuong} for phuong in phuongs]
        return JsonResponse({'wards': phuong_list})

#---------- Giỏ hàng
def cart(request):
     return render(request, 'pages/User/cart.html')

@login_required(login_url='/DangNhap/')
def add_to_cart(request):
    if request.method == 'POST':
        product_id = request.POST.get('variantId')
        quantity = int(request.POST.get('quantity', 1)) #mạc định số lượng là 1

        if product_id:
            product = get_object_or_404(SanPham, pk=product_id)
            cart, created = DonHang.objects.get_or_create(user=request.user, TrangThai='don_moi')
            cart_item, item_created = CT_DonHang.objects.get_or_create(gio_hang=cart, MaSP=product)

            #Nếu sản phẩm đã có trong giỏ thì cộng thêm vô
            cart_item.SoLuong = quantity if item_created else cart_item.SoLuong + quantity
            cart_item.save()
            cart.update_cart()

            return JsonResponse({
                'success': True,
                'product_name': product.TenSP,
                'product_price': product.GiaBan,
                'soLuong': cart_item.SoLuong,
                'tongTien': cart_item.ThanhTien,
                'anh': product.HinhAnh.url
            })
    return JsonResponse({'success': False, 'error': 'Invalid request'})

@login_required(login_url='/DangNhap/')
def listGioHang(request):
    user = request.user
    shipping_fee = 0  # Fixed shipping fee
    voucher = 0  # Voucher discount, modify if applicable
    city_lst = ThanhPho.objects.all()

     # Lấy tên phường, quận, thành phố từ các bảng liên kết
    ward_name = user.ward.TenPhuong if user.ward else "Không xác định"
    district_name = user.district.TenQuan if user.district else "Không xác định"
    city_name = user.city.TenTP if user.city else "Không xác định"
    user_address = ", ".join(filter(None, [user.street, ward_name, district_name, city_name]))   

    try:
        # Retrieve the existing cart for the user
        cart = DonHang.objects.get(user=user, TrangThai='don_moi')
        cart_items = CT_DonHang.objects.filter(gio_hang=cart)
        total_payment = cart.TongTien + shipping_fee + voucher
    except DonHang.DoesNotExist:
        # Create a new cart if none exists
        cart = DonHang.objects.create(
            user=request.user,
            TrangThai='don_moi',  # New cart, unconfirmed status
            user_name=user.full_name,
            user_address=user_address,
            user_phone=user.phone,
            TongTien=0,  # Initial total amount
        )
        cart_items = []  # Empty cart items
        total_payment = shipping_fee + voucher  # Base total with no items

    context = {
        'user_name': user.full_name,
        'user_phone': user.phone,
        'user_address': user_address,
        'DM_GH': cart_items,
        'TongTien': cart.TongTien,
        'ShippingFee': shipping_fee,
        'Voucher': voucher,
        'TotalPayment': total_payment,
        'city_lst': city_lst
    }

    return render(request, 'pages/User/cart.html', context)

def update_cart_total(cart):
   cart.update_cart()

def xoa_ctdh(request, ctdh_id):
    ctdh = get_object_or_404(CT_DonHang, pk=ctdh_id)
    cart = ctdh.gio_hang  # Lấy giỏ hàng từ chi tiết đơn hàng

    with transaction.atomic():
        ctdh.delete()  # Xoá sản phẩm khỏi giỏ hàng

        # Cập nhật lại tổng giỏ hàng
        update_cart_total(cart)

    return redirect('UserMember:GioHang')

def update_quantity(request, ctdh_id):
    if request.method == 'POST':
        # Lấy đối tượng CT_DonHang từ ID
        ctdh = get_object_or_404(CT_DonHang, pk=ctdh_id)
        cart = ctdh.gio_hang  # Giỏ hàng của sản phẩm này

        # Cập nhật số lượng sản phẩm
        new_quantity = int(request.POST.get('quantity'))
        ctdh.SoLuong = new_quantity
        ctdh.save()

        # Cập nhật tổng tiền giỏ hàng
        cart.update_cart()

        # Trả về kết quả cập nhật số lượng và tổng tiền mới
        return redirect('UserMember:GioHang')

    return JsonResponse({'error': 'Invalid request method'})

# Hàm áp dụng voucher
def apply_voucher(request):
    if request.method == 'POST':
        voucher_code = request.POST.get('voucher_code')
        total_amount = float(request.POST.get('total_amount'))
        # shipping_fee = float(request.POST.get('shipping_fee'))
        
        try:
            # Tìm voucher trong database
            voucher = Voucher.objects.get(MaVoucher=voucher_code)
            today = localtime(now()).date()

            # Kiểm tra ngày hết hạn và số lượng
            if not (voucher.NgayBatDau <= today <= voucher.NgayKetThuc):
                return JsonResponse({'error': "Voucher không còn hiệu lực."})
            if voucher.SoLuong <= 0:
                return JsonResponse({'error': "Voucher đã hết lượt sử dụng."})
            if total_amount < voucher.HanMucApDung:
                return JsonResponse({'error': "Tổng tiền chưa đủ để sử dụng voucher."})

            # Tính toán chiết khấu
            discount_amount = (voucher.PhanTramGiam / 100) * total_amount
            final_total = total_amount - discount_amount 

            return JsonResponse({
                'success': True,
                'discountAmount': f"{discount_amount:,.3f}",
                'finalTotal': f"{final_total:,.3f}"
            })

        except Voucher.DoesNotExist:
            return JsonResponse({'error': "Mã voucher không hợp lệ."})

    return JsonResponse({'error': 'Invalid request method'})

# Hàm xác nhận đơn hàng
def confirm_order(request):
    if request.method == 'POST':
        try:
            # Lấy giỏ hàng hiện tại
            cart = DonHang.objects.filter(user=request.user, TrangThai='don_moi').select_related('user').first()
            if not cart:
                return JsonResponse({'success': False, 'error': 'Giỏ hàng không tồn tại.'})

            # Lấy thông tin từ form
            user_name = request.POST.get('user_name', '').strip()
            user_phone = request.POST.get('user_phone', '').strip()
            address = request.POST.get('user_address', '').strip()
            voucher_code = request.POST.get('voucher_code', '').strip()
            print(f"Voucher Code: {voucher_code}")
            street = request.POST.get('street', '').strip()

            # Ánh xạ ID thành tên
            city_id = request.POST.get('city', '').strip()
            district_id = request.POST.get('district', '').strip()
            ward_id = request.POST.get('ward', '').strip()

            city_name = ThanhPho.objects.get(MaTP=city_id).TenTP if city_id else ''
            district_name = Quan.objects.get(MaQuan=district_id).TenQuan if district_id else ''
            ward_name = Phuong.objects.get(MaPhuong=ward_id).TenPhuong if ward_id else ''

            user_address = address if address else ', '.join(filter(None, [street, ward_name, district_name, city_name]))

            # Kiểm tra thông tin
            if not all([user_name, user_address, user_phone]):
                return JsonResponse({'success': False, 'error': 'Thông tin đặt hàng không đầy đủ.'})

            with transaction.atomic():
                # Cập nhật thông tin đơn hàng
                cart.user_name = user_name
                cart.user_address = user_address
                cart.user_phone = user_phone

                # Áp dụng voucher nếu có
                if voucher_code:
                    try:
                        voucher = Voucher.objects.get(MaVoucher=voucher_code)

                        # Kiểm tra nếu voucher còn đủ số lượng để áp dụng
                        if voucher.SoLuong > 0:
                            cart.voucher = voucher
                        
                            # Tính lại tổng tiền sau khi áp dụng voucher
                            discount_percentage = Decimal(voucher.PhanTramGiam) / 100
                            discount_amount = discount_percentage * Decimal(cart.TongTien)

                            cart.MucGiamGia = discount_amount
                            cart.TongTien -= discount_amount
                        
                            # Trừ số lượng voucher còn lại
                            voucher.SoLuong -= 1
                            voucher.save()
                        else:
                            return JsonResponse({'success': False, 'error': "Mã voucher đã hết lượt sử dụng."})
                        
                    except Voucher.DoesNotExist:
                        return JsonResponse({'success': False, 'error': "Mã voucher không hợp lệ."})

                # Cập nhật trạng thái đơn hàng và ngày dự kiến giao hàng
                cart.TrangThai = 'cho_xac_nhan'
                cart.NgayDat = localtime(now()).date()
                #cập nhật ngày dự kiến
                chi_tiet_don_hang = CT_DonHang.objects.filter(gio_hang=cart)            
                has_damaged_items = chi_tiet_don_hang.filter(MaSP__MucDoHu=True).exists() 
                # Tính toán NgayDuKien
                if has_damaged_items:
                    cart.NgayDuKien = cart.NgayDat + timedelta(days=1)
                else:
                    cart.NgayDuKien = cart.NgayDat + timedelta(days=2)

                 # Cập nhật số lượng sản phẩm trong kho
                chi_tiet_don_hang = CT_DonHang.objects.filter(gio_hang=cart)
                for chi_tiet in chi_tiet_don_hang:
                    san_pham = chi_tiet.MaSP
                    so_luong_dat = chi_tiet.SoLuong

                    ton_kho_queryset = TonKho.objects.filter(SanPham=san_pham).order_by('NgayHetHan')
                    total_stock = sum(item.SoLuongTon for item in ton_kho_queryset)

                   # Kiểm tra nếu tồn kho đủ đáp ứng
                    if total_stock >= so_luong_dat:
                        for ton_kho in ton_kho_queryset:
                            if so_luong_dat <= 0:
                                break

                            # Nếu tồn kho của lô đủ để trừ
                            if ton_kho.SoLuongTon >= so_luong_dat:
                                ton_kho.SoLuongTon -= so_luong_dat
                                ton_kho.save()
                                so_luong_dat = 0
                            else:
                                # Nếu không đủ thì trừ hết số lượng tồn của lô này, chuyển sang lô tiếp theo
                                so_luong_dat -= ton_kho.SoLuongTon
                                ton_kho.SoLuongTon = 0
                                ton_kho.save()
                    else:
                        return JsonResponse({'success': False, 'error': f'Sản phẩm {san_pham.TenSP} không đủ số lượng trong kho.'})
                
                # Lưu đơn hàng
                cart.save()

                # Tạo giỏ hàng mới
                DonHang.objects.create(user=request.user, TrangThai='don_moi', TongTien=0)

            # Trả về giao diện cập nhật
            return render(request, 'pages/User/cart.html')

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Invalid request method'})


#---------- Lịch sử mua hàng
@login_required(login_url='/DangNhap/')
def LSMuaHang(request):
    user = request.user
    orders = DonHang.objects.filter(user=user)

    # Lọc theo trạng thái
    status_filter = request.GET.get('status', 'all')  # Mặc định là 'all'
    if status_filter != 'all':
        orders = orders.filter(TrangThai=status_filter)

    # Chia đơn hàng theo trạng thái
    allorders =  orders.exclude(TrangThai='don_moi').order_by('-MaDH')
    pending_orders = orders.filter(TrangThai='cho_xac_nhan').order_by('-MaDH')
    processing_orders = orders.filter(TrangThai='dang_xu_ly').order_by('-MaDH')
    shipping_orders = orders.filter(TrangThai='dang_giao').order_by('-MaDH')
    completed_orders = orders.filter(TrangThai='da_hoan_thanh').order_by('-MaDH')

    # Kiểm tra đánh giá cho mỗi đơn hàng
    for order in completed_orders:
        order.has_reviewed = order.don_danh_gia.exists()

    return render(request, 'pages/User/LSMuaHang.html', {
        'allorders': allorders,
        'pending_orders': pending_orders,
        'processing_orders': processing_orders,
        'shipping_orders': shipping_orders,
        'completed_orders': completed_orders
    })
def get_ct_donhang(request, MaDH):
    # Lấy đối tượng đơn hàng
    don_hang = get_object_or_404(DonHang, pk=MaDH)
    
    # Lấy danh sách sản phẩm trong đơn hàng
    danh_sach_san_pham = []
    for chi_tiet in don_hang.ct_donhang.all():
        danh_sach_san_pham.append({
            'MaSP': chi_tiet.MaSP.MaSP,
            'TenSP': chi_tiet.MaSP.TenSP,
            'Gia': float(chi_tiet.MaSP.GiaBan),
            'SoLuong': chi_tiet.SoLuong,
            'ThanhTien': float(chi_tiet.MaSP.GiaBan) * chi_tiet.SoLuong,
        })

    # Chuẩn bị dữ liệu phản hồi
    ngay_nhan = don_hang.NgayNhan.strftime('%Y-%m-%d') if don_hang.NgayNhan else "Chưa nhận"
    trang_thai_dict = {
        'cho_xac_nhan': 'Chờ xác nhận',
        'dang_xu_ly': 'Đang xử lý',
        'dang_giao': 'Đang giao',
        'da_hoan_thanh': 'Đã hoàn thành',
    }
    
    trang_thai = trang_thai_dict.get(don_hang.TrangThai, 'Không xác định')
    data = {
        'MaDH': don_hang.MaDH,
        'TenKH': don_hang.user_name,
        'SdtKH': don_hang.user_phone,
        'NgayDat': don_hang.NgayDat.strftime('%Y-%m-%d'),
        'NgayNhan': ngay_nhan,
        'TongTien': don_hang.TongTien,
        'TrangThai': trang_thai,
        'SanPham': danh_sach_san_pham,
    }

    return JsonResponse(data)
@csrf_exempt
def update_order(request):
    if request.method == 'POST':
        order_id = request.POST.get('order_id')
        new_status = request.POST.get('new_status')

        try:
            # Lấy đơn hàng cần cập nhật
            order = DonHang.objects.get(MaDH=order_id)

            # Kiểm tra trạng thái đơn hàng
            if order.TrangThai == 'dang_giao' and new_status == 'da_hoan_thanh':
                order.TrangThai = 'da_hoan_thanh'
                order.NgayNhan = now() + timedelta(days=1)
                order.save()
                return JsonResponse({'success': True, 'message': 'Cập nhật trạng thái thành công.'})
            else:
                return JsonResponse({'success': False, 'error': 'Không thể thay đổi trạng thái này.'})
        except DonHang.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Đơn hàng không tồn tại.'})

    return JsonResponse({'success': False, 'error': 'Invalid request method'})

#---------- Sản phẩm
def loaiSP(request):
    loai_san_pham_list = Loai.objects.all()
    return {'loaiSanPhamList': loai_san_pham_list}

def chiTietSP(request, product_id):
    san_pham = get_object_or_404(SanPham, MaSP=product_id)
    loaiSP = SanPham.objects.filter(MaLoai_id=san_pham.MaLoai).exclude(MaSP=product_id)[:4]
    return render(request, 'pages/User/chiTietSP.html', {'san_pham': san_pham, 'loai_sp': loaiSP})

def DSSanPham(request):
    # Lấy tham số từ URL
    keyword = request.GET.get("q", "")  # Tìm kiếm
    sort = request.GET.get('sort', '')  # Sắp xếp
    ma_loai = request.GET.get('loai')  # Lọc theo loại

    # Danh sách sản phẩm ban đầu
    product_list = SanPham.objects.all()

    # Tìm kiếm
    if keyword:
        product_list = product_list.filter(TenSP__icontains=keyword)

    # Lọc theo loại sản phẩm (kiểm tra giá trị ma_loai)
    if ma_loai and ma_loai != 'None':
        loai_sp = Loai.objects.get(MaLoai=ma_loai)
        product_list = product_list.filter(MaLoai=ma_loai)
        ten_loai = loai_sp.TenLoai
    else:
        ten_loai = "Tất cả sản phẩm"

    # Sắp xếp sau khi đã tìm kiếm và lọc
    if sort == 'GiaTang':
        product_list = product_list.order_by('GiaBan')
    elif sort == 'GiaGiam':
        product_list = product_list.order_by('-GiaBan')

    # Phân trang
    paginator = Paginator(product_list, 6)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Lấy danh sách loại sản phẩm
    loai_san_pham_list = Loai.objects.all()

    # Trả về kết quả
    return render(request, 'pages/User/DSSanPham.html', {
        'dsSanPham': page_obj,
        'ten_loai': ten_loai,
        'ma_loai': ma_loai,
        'sort': sort,
        'loai_san_pham_list': loai_san_pham_list,
        'keyword': keyword,
    })

#----------------------------------------A D M I N -----------------------------------------------
# Hàm kiểm tra xem người dùng có phải admin không
def is_admin(user):
    return user.is_admin
def is_admin_or_staff(user):
    return user.is_admin or user.is_staff
def is_admin_or_manager(user):
    return user.is_admin or user.is_manager
def is_admin_or_staff_or_manager(user):
    return user.is_admin or user.is_staff or user.is_manager


def qlHome(request):
    return render(request, 'pages/Admin/qlHome.html')

#----------QLDonHang
@login_required(login_url='/DangNhap/')  # Đảm bảo người dùng đã đăng nhập
@user_passes_test(is_admin_or_staff_or_manager, login_url='/DangNhap/')
def QLDonHang(request):
    orders = DonHang.objects.exclude(TrangThai='don_moi').order_by('-MaDH')
    return render(request, 'pages/Admin/QLDonHang.html', {'orders': orders})

@require_POST
def update_order_status(request):
    order_id = request.POST.get('order_id')
    new_status = request.POST.get('new_status')

    try:
        # Lấy đơn hàng cần cập nhật
        order = DonHang.objects.get(MaDH=order_id)

        # Kiểm tra trạng thái hiện tại trước khi thay đổi
        if order.TrangThai == 'cho_xac_nhan' and new_status == 'dang_xu_ly':
            order.TrangThai = 'dang_xu_ly'
            order.save()
            return JsonResponse({'success': True, 'message': 'Cập nhật trạng thái thành công.'})

        elif order.TrangThai == 'dang_xu_ly' and new_status == 'dang_giao':
            order.TrangThai = 'dang_giao'
            order.save()
            return JsonResponse({'success': True, 'message': 'Cập nhật trạng thái thành công.'})

        else:
            return JsonResponse({'success': False, 'error': 'Trạng thái không hợp lệ hoặc không thể thay đổi trạng thái ở giai đoạn này.'})
    except DonHang.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Đơn hàng không tồn tại.'})

#----------QLLoai 
@login_required(login_url='/DangNhap/')
@user_passes_test(is_admin_or_manager, login_url='/DangNhap/')
def QLLoai(request):
    dsLoaiSP = Loai.objects.all()
    if request.method == 'POST':
        ma_loai = request.POST.get('MaLoai')
        
        if ma_loai:  # Nếu có MaLoai, tức là đang sửa
            loai = get_object_or_404(Loai, pk=ma_loai)
            form = LoaiForm(request.POST, instance=loai)
        else:  # Không có MaLoai tức là thêm mới
            form = LoaiForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('UserMember:QLLoai')  # Sau khi thêm/sửa, quay lại danh sách
    
    return render(request, 'pages/Admin/QLLoai.html', {'dsLoaiSP': dsLoaiSP})
def them_loai(request):
    if request.method == 'POST':
        form = LoaiForm(request.POST)
        if form.is_valid():
            with transaction.atomic():
                # Lấy giá trị MaLoai cao nhất
                max_ma_loai = Loai.objects.aggregate(max_ma_loai=Max('MaLoai'))['max_ma_loai']
                new_ma_loai = max_ma_loai + 1 if max_ma_loai else 1
                loai = form.save(commit=False)
                loai.MaLoai = new_ma_loai
                loai.save()
            return redirect('UserMember:QLLoai')
    else:
        form = LoaiForm()
    
    return render(request, 'pages/Admin/them_sua_loai.html', {'form': form, 'title': 'Thêm Loại Sản Phẩm'})
def sua_loai(request, loai_id):
    loai = get_object_or_404(Loai, pk=loai_id)
    if request.method == 'POST':
        form = LoaiForm(request.POST, instance=loai)
        if form.is_valid():
            form.save()
            return redirect('UserMember:QLLoai') 
    else:
        form = LoaiForm(instance=loai)
    
    return render(request, 'pages/Admin/them_sua_loai.html', {'form': form, 'title': 'Sửa Loại Sản Phẩm'})
def xoa_loai(request, loai_id):
    loai = get_object_or_404(Loai, pk=loai_id)
    with transaction.atomic():
        loai.delete()  # Chỉ xóa loại sản phẩm mà không thay đổi thứ tự MaLoai
    return redirect('UserMember:QLLoai') 

#----------QLSanPham
@login_required(login_url='/DangNhap/')
@user_passes_test(is_admin_or_staff_or_manager, login_url='/DangNhap/')
def QLSanPham(request):
    # Lấy từ khóa tìm kiếm và loại sản phẩm
    keyword = request.GET.get("q", "").strip()
    loai_id = request.GET.get('loai', None)

    # Khởi tạo danh sách sản phẩm
    dsSanPham = SanPham.objects.filter(IsDelete__isnull=True)

    # Lọc theo loại (nếu có)
    selected_loai = None
    if loai_id:
        dsSanPham = dsSanPham.filter(MaLoai__MaLoai=loai_id)
        selected_loai = Loai.objects.filter(MaLoai=loai_id).first()

    # Tìm kiếm theo từ khóa (nếu có)
    if keyword:
        dsSanPham = dsSanPham.filter(TenSP__icontains=keyword)

    # Lấy danh sách loại sản phẩm
    dsLoai = Loai.objects.all()

    return render(request, 'pages/Admin/QLSanPham.html', {
        'dsSanPham': dsSanPham,
        'dsLoai': dsLoai,
        'selected_loai': selected_loai,
        'keyword': keyword,
    })

def them_san_pham(request):
    if request.method == 'POST':
        form = SanPhamForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('UserMember:QLSanPham')
    else:
        form = SanPhamForm()
    
    return render(request, 'pages/Admin/ThemSanPham.html', {'form': form, 'title': 'Thêm Sản Phẩm'})
def sua_san_pham(request, sanpham_id):
    san_pham = get_object_or_404(SanPham, pk=sanpham_id)
    if request.method == 'POST':
        form = SanPhamForm(request.POST, request.FILES, instance=san_pham)
        if form.is_valid():
            san_pham.cap_nhat_so_luong()
            form.save()
            return redirect('UserMember:QLSanPham')
    else:
        form = SanPhamForm(instance=san_pham)
    
    return render(request, 'pages/Admin/SuaSanPham.html', {'form': form, 'title': 'Sửa Sản Phẩm'})    
def xoa_san_pham(request, sanpham_id):
    sanpham = get_object_or_404(SanPham, MaSP=sanpham_id)  # Lấy sản phẩm theo MaSP
    if request.method == 'POST':
        sanpham.delete()  # Xóa sản phẩm
        return redirect('UserMember:QLSanPham')  # Chuyển hướng về trang danh sách sản phẩm
    return render(request, 'pages/Admin/xoa_confirm.html', {'sanpham': sanpham})
def get_san_pham(request, MaSP):
    san_pham = get_object_or_404(SanPham, pk=MaSP)
    # Truy vấn TonKho với trường NgayHetHan
    tonkho_list = TonKho.objects.filter(SanPham=san_pham).order_by('NgayHetHan')
    hsd_info = [
        {"SoLuongTon": tk.SoLuongTon, "NgayHetHan": tk.NgayHetHan} for tk in tonkho_list if tk.SoLuongTon > 0
    ]
    print("HSD Info:", hsd_info)
    data = {
        'MaSP': san_pham.MaSP,
        'TenSP': san_pham.TenSP,
        'TrongLuong': san_pham.TrongLuong,
        'NguonGoc': san_pham.NguonGoc,
        'DVT': {
            'TenDonVi': san_pham.DVT.dvt,
        }, 
        'SoLuongHienTai': san_pham.SoLuongHienTai,
        'GiaBan': san_pham.GiaBan,
        'BaoQuan': san_pham.BaoQuan,
        'HinhAnh': san_pham.HinhAnh.url if san_pham.HinhAnh else '',
        'TinhTrang': san_pham.TinhTrang,
        'MoTa': san_pham.MoTa,
        'MaLoai': {'TenLoai': san_pham.MaLoai.TenLoai},
        'HSD': hsd_info,  # Trả về danh sách hạn sử dụng
    }
    return JsonResponse(data)
#---------------------
@receiver(post_save, sender=TonKho)
def cap_nhat_so_luong_ton_kho(sender, instance, created, **kwargs):
    instance.SanPham.cap_nhat_so_luong()

def cap_nhat_so_luong(self):
    try:
        sl_hien_tai = TonKho.objects.filter(SanPham=self).aggregate(
            tong_ton=Sum('SoLuongTon')
        )['tong_ton'] or 0

        self.SoLuongHienTai = sl_hien_tai
        self.save()
    except Exception as e:
        print(f"Lỗi khi cập nhật số lượng cho sản phẩm {self.TenSP}: {e}")

#----------San Pham Hư
@login_required(login_url='/DangNhap/')
@user_passes_test(is_admin_or_staff, login_url='/DangNhap/')
def QLSanPhamHu(request):
    keyword = request.GET.get("q", "").strip()
    loai_id = request.GET.get('loai', None)

    dsSanPhamhu = SanPhamHu.objects.filter(IsDelete=False)
    selected_loai = None

    if loai_id:
        dsSanPhamhu = SanPhamHu.objects.filter(MaSP__MaLoai=loai_id)
        selected_loai = Loai.objects.get(MaLoai=loai_id)

    # Tìm kiếm theo từ khóa (nếu có)
    if keyword:
        dsSanPhamhu = dsSanPhamhu.filter(MaSP__TenSP__icontains=keyword)
   
    dsLoai = Loai.objects.all()

    return render(request, 'pages/Admin/QLSanPhamHu.html', {
        'dsSanPhamhu': dsSanPhamhu,
        'dsLoai': dsLoai,
        'selected_loai': selected_loai,
        'keyword': keyword
    })
def them_sanpham_hu(request):
    if request.method == 'POST':
        form = SanPhamHuForm(request.POST, request.FILES)
        if form.is_valid():
            san_pham_hu = form.save(commit=False)
            so_luong_hu = san_pham_hu.SoLuong
            san_pham = san_pham_hu.MaSP

            # Lấy danh sách tồn kho của sản phẩm theo ngày hết hạn gần nhất
            ton_kho_queryset = TonKho.objects.filter(SanPham=san_pham).order_by('NgayHetHan')

            total_stock = sum(item.SoLuongTon for item in ton_kho_queryset)
            if total_stock >= so_luong_hu:
                for ton_kho in ton_kho_queryset:
                    if so_luong_hu <= 0:
                        break

                    if ton_kho.SoLuongTon >= so_luong_hu:
                        # Trừ số lượng trong kho hiện tại
                        ton_kho.SoLuongTon -= so_luong_hu
                        ton_kho.save()
                        so_luong_hu = 0
                    else:
                        # Trừ hết số lượng tồn của lô này, chuyển sang lô tiếp theo
                        so_luong_hu -= ton_kho.SoLuongTon
                        ton_kho.SoLuongTon = 0
                        ton_kho.save()

                # Lưu sản phẩm hư
                san_pham_hu.NgayHu = now()
                san_pham_hu.save()
                messages.success(request, 'Thêm sản phẩm hư thành công.')
                return redirect('UserMember:QLSanPhamHu')
            else:
                messages.error(request, 'Số lượng hư vượt quá số lượng tồn kho của sản phẩm.')
        else:
            messages.error(request, 'Có lỗi trong quá trình thêm sản phẩm hư. Vui lòng thử lại.')
    else:
        form = SanPhamHuForm()

    return render(request, 'pages/Admin/ThemSanPhamHu.html', {
        'form': form,
        'title': 'Thêm Sản Phẩm Hư',
        'button_text': 'Thêm',
        'mode': 'add'
    })
def sua_sanpham_hu(request, ma_sp_hu):
    sanpham_hu = get_object_or_404(SanPhamHu, pk=ma_sp_hu)
    old_so_luong_hu = sanpham_hu.SoLuong

    if request.method == 'POST':
        form = SanPhamHuForm(request.POST, request.FILES, instance=sanpham_hu)
        if form.is_valid():
            san_pham = sanpham_hu.MaSP
            new_so_luong_hu = form.cleaned_data['SoLuong']

            # Kiểm tra nếu số lượng hư vượt quá số lượng hiện tại
            if new_so_luong_hu > san_pham.SoLuongHienTai + old_so_luong_hu:
                messages.error(request, 'Số lượng hư vượt quá số lượng hiện tại của sản phẩm.')
            else:
                san_pham.SoLuongHienTai += (old_so_luong_hu - new_so_luong_hu)
                san_pham.save()

                form.save()
                messages.success(request, 'Sửa sản phẩm hư thành công.')
                return redirect('UserMember:QLSanPhamHu')
        else:
            messages.error(request, 'Dữ liệu không hợp lệ. Vui lòng kiểm tra lại.')
    else:
        form = SanPhamHuForm(instance=sanpham_hu)

    return render(request, 'pages/Admin/ThemSanPhamHu.html', {
        'form': form,
        'title': 'Sửa Sản Phẩm Hư',
        'button_text': 'Cập nhật', 
        'mode': 'edit'
    })
def xoa_sanpham_hu(request, ma_sp_hu):
    sanpham_hu = get_object_or_404(SanPhamHu, pk=ma_sp_hu)
    if request.method == 'POST':
        sanpham_hu.delete()  # Xóa sản phẩm hư
        messages.success(request, 'Sản phẩm hư đã được xóa thành công.')
        return redirect('UserMember:QLSanPhamHu')  # Quay lại danh sách sản phẩm hư
    return redirect('UserMember:QLSanPhamHu')  # Nếu không phải POST, quay lại danh sách

def xuat_excel_san_pham_hu(request):
    # Lấy danh sách sản phẩm hư từ database
    ds_san_pham_hu = SanPhamHu.objects.filter(IsDelete=False).select_related('MaSP')

    # Tạo workbook và worksheet mới
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách sản phẩm hư"

    # Tiêu đề cột
    headers = ['STT', 'Mã sản phẩm hư', 'Tên sản phẩm', 'Số lượng', 'Ngày hư', 'Lý do', 'Hình ảnh']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Ghi dữ liệu từng dòng
    for idx, sp in enumerate(ds_san_pham_hu, start=1):
        ws.cell(row=idx + 1, column=1, value=idx)  # STT
        ws.cell(row=idx + 1, column=2, value=sp.MaSP_Hu)  # Mã sản phẩm hư
        ws.cell(row=idx + 1, column=3, value=sp.MaSP.TenSP if sp.MaSP else "(Không xác định)")  # Tên sản phẩm
        ws.cell(row=idx + 1, column=4, value=sp.SoLuong)  # Số lượng
        ws.cell(row=idx + 1, column=5, value=sp.NgayHu.strftime('%d/%m/%Y') if sp.NgayHu else "")  # Ngày hư
        ws.cell(row=idx + 1, column=6, value=sp.LyDo)  # Lý do
        ws.cell(row=idx + 1, column=7, value=sp.HinhAnh.url if sp.HinhAnh else "Không có ảnh")  # Hình ảnh

    # Định dạng cột (căn chỉnh và điều chỉnh độ rộng)
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    # Tạo response trả về file Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="danh_sach_san_pham_hu.xlsx"'
    wb.save(response)

    return response
#----------QLTaiKhoan
User = get_user_model()
@login_required(login_url='/DangNhap/')
@user_passes_test(is_admin_or_manager, login_url='/DangNhap/')
def QLTaiKhoanKhachHang(request):
    dsTaiKhoanKhachHang = User.objects.filter(is_user=True).order_by('id')
    for index, user in enumerate(dsTaiKhoanKhachHang, start=1):
        user.stt_khachhang = index 
    return render(request, 'pages/Admin/QLTaiKhoanKhachHang.html', {'dsTaiKhoanKhachHang': dsTaiKhoanKhachHang})
def get_user_details(request, user_id):
    user = get_object_or_404(User, pk=user_id)
    user_data = {
        'id': user.id,
        'username': user.username,
        'email': user.email,
    }
    return JsonResponse(user_data)
def sua_taikhoan_khachhang(request, user_id):
    user = get_object_or_404(User, pk=user_id)
    if request.method == 'POST':
        form = UserForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            update_stt_khachhang()
            return redirect('UserMember:QLTaiKhoanKhachHang')
    else:
        form = UserForm(instance=user)
    return render(request, 'pages/Admin/sua_taikhoan_khachhang.html', {'form': form, 'title': 'Sửa Khách Hàng'})

def deactivate_user(request, user_id):
    user = get_object_or_404(CustomUser, pk=user_id)
    if user.is_active:
        user.is_active = False
        user.save()
    return redirect('UserMember:QLTaiKhoanKhachHang')

def toggle_user_status(request, user_id):
    # Lấy thông tin người dùng
    user = get_object_or_404(CustomUser, pk=user_id)
    action = request.GET.get('action', '')

    if action == 'deactivate' and user.is_active:
        # Vô hiệu hóa tài khoản
        user.is_active = False
        user.save()

        # Gửi email thông báo vô hiệu hóa
        current_site = get_current_site(request)
        mail_subject = 'Thông báo vô hiệu hóa tài khoản'
        message = render_to_string('pages/User/Email_VoHieuHoa.html', {
            'user': user,
            'domain': current_site.domain,
        })
        send_mail(
            mail_subject,
            message,
            'ngtiendung012@gmail.com',  # Email gửi
            [user.email],  # Email nhận
            fail_silently=False,
        )

        # Thêm thông báo
        messages.success(request, f"Tài khoản của {user.username} đã bị vô hiệu hóa và email đã được gửi.")

    elif action == 'activate' and not user.is_active:
        # Kích hoạt lại tài khoản
        user.is_active = True
        user.save()

        # Gửi email thông báo kích hoạt lại
        current_site = get_current_site(request)
        mail_subject = 'Thông báo kích hoạt tài khoản'
        message = render_to_string('pages/User/Email_KichHoat.html', {
            'user': user,
            'domain': current_site.domain,
        })
        send_mail(
            mail_subject,
            message,
            'ngtiendung012@gmail.com',  # Email gửi
            [user.email],  # Email nhận
            fail_silently=False,
        )

        # Thêm thông báo
        messages.success(request, f"Tài khoản của {user.username} đã được kích hoạt và email đã được gửi.")

    return redirect('UserMember:QLTaiKhoanKhachHang')

def update_stt_khachhang():
    user_list = User.objects.filter(is_user=True).order_by('id')  # Sắp xếp theo id người dùng
    for index, user in enumerate(user_list, start=1):
        user.stt_khachhang = index 

#---------TaiKhoanNhanVien
@login_required(login_url='/DangNhap/')
@user_passes_test(is_admin, login_url='/DangNhap/')
def QLTaiKhoanNhanVien(request):
    city_lst = ThanhPho.objects.all() 
    dsTaiKhoanNhanVien = CustomUser.objects.filter(is_staff=True).order_by('id')
    for index, nhanvien in enumerate(dsTaiKhoanNhanVien, start=1):
        nhanvien.stt_nhanvien = index
    return render(request, 'pages/Admin/QLTaiKhoanNhanVien.html', {'dsTaiKhoanNhanVien': dsTaiKhoanNhanVien, 'city_lst': city_lst})

def them_nhanvien(request):
    if request.method == 'POST':
        # Lấy dữ liệu từ form
        full_name = request.POST.get('full_name')
        username = request.POST.get('username')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        street = request.POST.get('street')
        city_id = request.POST.get('city')
        district_id = request.POST.get('district')
        ward_id = request.POST.get('ward')
        password = request.POST.get('password')

        # Kiểm tra dữ liệu đầu vào
        errors = {}
        if not all([full_name, username, email, street, city_id, district_id, ward_id]):
            errors['required'] = "Vui lòng điền đầy đủ các thông tin bắt buộc."
        elif CustomUser.objects.filter(username=username).exists():
            errors['username'] = "Tên đăng nhập đã tồn tại."
        elif CustomUser.objects.filter(email=email).exists():
            errors['email'] = "Email đã tồn tại."

        if errors:
            return JsonResponse({'success': False, 'errors': errors}, status=400)

        try:
            # Tìm các đối tượng liên quan
            city = get_object_or_404(ThanhPho, pk=city_id)
            district = get_object_or_404(Quan, pk=district_id)
            ward = get_object_or_404(Phuong, pk=ward_id)

            # Tạo nhân viên mới
            CustomUser.objects.create_staff(
                email=email,
                username=username,
                phone=phone,
                full_name=full_name,
                street=street,
                city=city,
                district=district,
                ward=ward,
                password=make_password(password) if password else None
            )
            return JsonResponse({'success': True, 'message': "Thêm nhân viên thành công!"}, status=200)
        except Exception as e:
            return JsonResponse({'success': False, 'errors': {'server': f"Lỗi hệ thống: {e}"}}, status=500)

def sua_nhanvien(request, user_id):
    user = get_object_or_404(CustomUser, pk=user_id)
    if request.method == 'POST':
        # Lấy dữ liệu từ form
        full_name = request.POST.get('full_name')
        username = request.POST.get('username')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        street = request.POST.get('street')
        city_id = request.POST.get('city')
        district_id = request.POST.get('district')
        ward_id = request.POST.get('ward')
        password = request.POST.get('password')

        # Kiểm tra dữ liệu đầu vào
        errors = {}
        if not all([full_name, username, email, street, city_id, district_id, ward_id]):
            errors['required'] = "Vui lòng điền đầy đủ các thông tin bắt buộc."
        elif CustomUser.objects.filter(username=username).exclude(pk=user_id).exists():
            errors['username'] = "Tên đăng nhập đã tồn tại."
        elif CustomUser.objects.filter(email=email).exclude(pk=user_id).exists():
            errors['email'] = "Email đã tồn tại."

        if errors:
            return JsonResponse({'success': False, 'errors': errors}, status=400)

        try:
            # Tìm các đối tượng liên quan
            city = get_object_or_404(ThanhPho, pk=city_id)
            district = get_object_or_404(Quan, pk=district_id)
            ward = get_object_or_404(Phuong, pk=ward_id)

            # Cập nhật thông tin nhân viên
            user.full_name = full_name
            user.username = username
            user.email = email
            user.phone = phone
            user.street = street
            user.city = city
            user.district = district
            user.ward = ward
            if password:
                user.password = make_password(password)
            user.save()

            return JsonResponse({'success': True, 'message': "Cập nhật nhân viên thành công!"}, status=200)
        except Exception as e:
            return JsonResponse({'success': False, 'errors': {'server': f"Lỗi hệ thống: {e}"}}, status=500)

def xoa_nhanvien(request, user_id):
    user = get_object_or_404(CustomUser, pk=user_id)
    try:
        user.delete()
        messages.success(request, "Nhân viên đã được xóa thành công.")
    except Exception as e:
        messages.error(request, f"Lỗi khi xóa nhân viên: {e}")
    return redirect('UserMember:QLTaiKhoanNhanVien')

def check_unique_field(request):
    field = request.GET.get('field')
    value = request.GET.get('value')
    is_unique = True
    message = ""

    if field not in ['username', 'email']:
        return JsonResponse({'error': 'Trường không hợp lệ.'}, status=400)

    if field == 'username' and CustomUser.objects.filter(username=value).exists():
        is_unique = False
        message = "Tên đăng nhập đã tồn tại."
    elif field == 'email' and CustomUser.objects.filter(email=value).exists():
        is_unique = False
        message = "Email đã tồn tại."

    return JsonResponse({'is_unique': is_unique, 'message': message}, status=200)
 
#----------------------------------------------------------------------------------------------------------
def cap_quyen_nhanvien(request, user_id):
    nhanvien = get_object_or_404(CustomUser, pk=user_id, is_staff=True)
    nhanvien.is_manager = not nhanvien.is_manager
    nhanvien.save()
    if nhanvien.is_manager:
        messages.success(request, f"Đã cấp quyền quản lý toàn bộ cho {nhanvien.full_name}.")
    else:
        messages.warning(request, f"Đã thu hồi quyền quản lý toàn bộ của {nhanvien.full_name}.")

    return redirect('UserMember:QLTaiKhoanNhanVien')  # Quay lại trang danh sách nhân viên

#---------QuenMatKhau
def recover(request):
    if request.method == 'POST':
        email = request.POST['Email']
        user = CustomUser.objects.filter(email=email).first()
        if user:
            token = default_token_generator.make_token(user)
            uid = urlsafe_base64_encode(force_bytes(user.pk))
            current_site = get_current_site(request)
            mail_subject = 'Lấy lại mật khẩu'
            message = render_to_string('pages/User/reset_password_email.html', {
                'user': user,
                'domain': current_site.domain,
                'uid': uid,
                'token': token,
            })
            send_mail(mail_subject, message, 'Ngtiendung012@gmail.com', [email])
            return JsonResponse({'message': 'Vui lòng kiểm tra email của bạn để lấy lại mật khẩu.'}, status=200)
        else:
            return JsonResponse({'message': 'Không tìm thấy người dùng với email này.'}, status=404)
    return JsonResponse({'message': 'Yêu cầu không hợp lệ.'}, status=400)

from .forms import PasswordResetForm

def password_reset_confirm(request, uidb64, token):
    UserModel = get_user_model()

    try:
        # Giải mã UID và lấy user
        uid = urlsafe_base64_decode(uidb64).decode()
        user = UserModel._default_manager.get(pk=uid)
    except (TypeError, ValueError, OverflowError, UserModel.DoesNotExist):
        user = None

    if user is not None and default_token_generator.check_token(user, token):
        if request.method == 'POST':
            form = PasswordResetForm(request.POST)
            if form.is_valid():
                # Cập nhật mật khẩu mới
                new_password = form.cleaned_data.get('password')
                user.set_password(new_password)
                user.save()
                return redirect('UserMember:DangNhap')
        else:
            form = PasswordResetForm(initial={'email': user.email})  # Điền sẵn email vào form
        return render(request, 'pages/User/resetMatKhau.html', {'form': form})
    else:
        # Token không hợp lệ hoặc đã hết hạn
        return redirect('UserMember:Recover')


#----------QLTinTuc 
@login_required(login_url='/DangNhap/')
@user_passes_test(is_admin, login_url='/DangNhap/')
def QLTinTuc(request):
    return render(request, 'pages/Admin/QLTinTuc.html')

#---------- QLVoucher
@login_required(login_url='/DangNhap/')
@user_passes_test(is_admin, login_url='/DangNhap/')
def QLVoucher(request):
    today = now().date()
    dsVoucher = Voucher.objects.all()
    
    if request.method == 'POST':
        ma_voucher = request.POST.get('MaVoucher')
        
        # Nếu có MaVoucher, tức là đang sửa, nếu không có, là thêm mới
        if ma_voucher:  
            voucher = get_object_or_404(Voucher, pk=ma_voucher)
            form = VoucherForm(request.POST, instance=voucher)
        else:
            form = VoucherForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('UserMember:QLVoucher') 
    else:
        form = VoucherForm() 

    return render(request, 'pages/Admin/QLVoucher.html', {
        'dsVoucher': dsVoucher,
        'form': form,
        'title': 'Quản lý Voucher',
        'today': today
    })
def them_voucher(request):
    if request.method == 'POST':
        form = VoucherForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('UserMember:QLVoucher')
        else:
            print(form.errors)  # In ra lỗi nếu form không hợp lệ
    else:
        form = VoucherForm()
    return render(request, 'pages/Admin/them_sua_voucher.html', {'form': form, 'title': 'Thêm Voucher'})
def sua_voucher(request, MaVoucher):
    voucher = get_object_or_404(Voucher, pk=MaVoucher)
    if request.method == 'POST':
        form = VoucherForm(request.POST, instance=voucher)
        if form.is_valid():
            form.save()
            return redirect('UserMember:QLVoucher')
    else:
        form = VoucherForm(instance=voucher)
    return render(request, 'pages/Admin/them_sua_voucher.html', {'form': form, 'title': 'Sửa Voucher'})
def xoa_voucher(request, MaVoucher):
    voucher = get_object_or_404(Voucher, pk=MaVoucher)
    voucher.delete()
    messages.success(request, 'Voucher đã được xóa thành công.')
    return redirect('UserMember:QLVoucher')

#------------QL NCC
@login_required(login_url='/DangNhap/')
@user_passes_test(is_admin, login_url='/DangNhap/')
def QLncc(request):
    dsncc = NhaCungCap.objects.filter(IsDelete__isnull=True)  # Filter for suppliers not deleted
    return render(request, 'pages/Admin/QLNCC.html', {'dsncc': dsncc})

def them_ncc(request):
    if request.method == 'POST':
        form = NhaCungCapForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Nhà cung cấp đã được thêm thành công.')
            return redirect('UserMember:QLNCC') 
        else:
            messages.error(request, 'Có lỗi xảy ra. Vui lòng kiểm tra lại thông tin.')
    else:
        form = NhaCungCapForm()
    return render(request, 'pages/Admin/them_ncc.html', {'form': form, 'title': 'Thêm Nhà Cung Cấp'})
def sua_ncc(request, MaNCC):
    ncc = get_object_or_404(NhaCungCap, pk=MaNCC)
    if request.method == 'POST':
        form = NhaCungCapForm(request.POST, instance=ncc)
        if form.is_valid():
            form.save()
            messages.success(request, 'Thông tin nhà cung cấp đã được cập nhật.')
            return redirect('UserMember:QLNCC')
        else:
            messages.error(request, 'Có lỗi xảy ra. Vui lòng kiểm tra lại thông tin.')
    else:
        form = NhaCungCapForm(instance=ncc)
    return render(request, 'pages/Admin/sua_ncc.html', {'form': form, 'title': 'Sửa Nhà Cung Cấp'})
def xoa_ncc(request, MaNCC):
    ncc = get_object_or_404(NhaCungCap, pk=MaNCC)
    ncc.delete()  # Xóa nhà cung cấp
    messages.success(request, 'Nhà cung cấp đã được xóa thành công.')
    return redirect('UserMember:QLNCC')

#------Phiếu nhập 
@login_required(login_url='/DangNhap/')
@user_passes_test(is_admin, login_url='/DangNhap/')
def QLPhieunhap(request):
    dsPhieuNhap = PhieuNhap.objects.filter(IsDelete__isnull=True)  
    nha_cung_cap_list = NhaCungCap.objects.all()
    phieu_nhap_form = PhieuNhapForm()
    dskho_info = Kho.objects.all().values('MaKho', 'TenKho') 
    dskho_info_list = list(dskho_info)
    return render(request, 'pages/Admin/QLPhieuNhap.html', {
        'dsPhieuNhap': dsPhieuNhap,  # Danh sách phiếu nhập
        'nha_cung_cap_list': nha_cung_cap_list,  # Danh sách nhà cung cấp
        'phieu_nhap_form': phieu_nhap_form,  # Form phiếu nhập
        'dskho_info': dskho_info_list
    })

def them_phieu_nhap(request):
    if request.method == 'POST':
        phieu_nhap_form = PhieuNhapForm(request.POST)
        if phieu_nhap_form.is_valid():
            phieu_nhap = phieu_nhap_form.save(commit=False)  
            phieu_nhap.TongSoLuong = 0 
            phieu_nhap.TongTienNhap = 0  
            phieu_nhap.save() 
            
            messages.success(request, "Phiếu nhập mới đã được thêm thành công!")
            return redirect('UserMember:QLPhieuNhap') 
        else:
            messages.error(request, "Có lỗi xảy ra khi thêm phiếu nhập.")
    else:
        phieu_nhap_form = PhieuNhapForm()

    return render(request, 'pages/Admin/them_phieu_nhap.html', {
        'phieu_nhap_form': phieu_nhap_form,
        'title': 'Thêm Phiếu Nhập Mới'
    })
def sua_phieu_nhap(request, MaPhieuNhap):
    phieu_nhap = get_object_or_404(PhieuNhap, MaPhieuNhap=MaPhieuNhap)

    if request.method == 'POST':
        phieu_nhap_form = PhieuNhapForm(request.POST, instance=phieu_nhap)
        
        if phieu_nhap_form.is_valid():
            phieu_nhap_form.save()  # Lưu lại thông tin tổng quát
            messages.success(request, "Thông tin phiếu nhập đã được cập nhật thành công!")
            return redirect('UserMember:QLPhieuNhap')
        else:
            messages.error(request, "Có lỗi xảy ra khi cập nhật phiếu nhập.")

    else:
        phieu_nhap_form = PhieuNhapForm(instance=phieu_nhap)
    
    return render(request, 'pages/Admin/sua_phieu_nhap.html', {
        'phieu_nhap_form': phieu_nhap_form,
        'title': 'Sửa Thông Tin Phiếu Nhập'
    })
def xoa_phieu_nhap(request, MaPhieuNhap):
    phieu_nhap = get_object_or_404(PhieuNhap, MaPhieuNhap=MaPhieuNhap)
    if phieu_nhap.TrangThai == 'da_nhap':
        for chi_tiet in phieu_nhap.ct_phieunhap.all():
            san_pham = chi_tiet.SanPham
            san_pham.SoLuongHienTai -= chi_tiet.SoLuongNhap  # Trừ số lượng đã nhập
            san_pham.save()
    phieu_nhap.delete()
    messages.success(request, "Phiếu nhập đã được xóa thành công.")
    return redirect('UserMember:QLPhieuNhap')

def XemChiTietPhieuNhap(request, MaPhieuNhap):
    phieu_nhap = get_object_or_404(PhieuNhap, MaPhieuNhap=MaPhieuNhap)
    chi_tiet_phieu_nhap = CT_PhieuNhap.objects.filter(PhieuNhap=phieu_nhap)
    loai_san_pham = Loai.objects.all()
    san_pham = SanPham.objects.all()
    for sp in chi_tiet_phieu_nhap:
        if not sp.MaCTPN: 
            sp.MaCTPN = None  
    context = {
        'phieu_nhap': phieu_nhap,
        'chi_tiet_phieu_nhap': chi_tiet_phieu_nhap,
        'loai_san_pham': loai_san_pham,
        'san_pham': san_pham,
        'title': 'Chi Tiết Phiếu Nhập'
    }
    return render(request, 'pages/Admin/XemChiTietPhieuNhap.html', context)

def get_san_pham_by_loai(request, loai_id):
    san_phams = SanPham.objects.filter(MaLoai__MaLoai=loai_id)
    data = []
    for sp in san_phams:
        data.append({
            'MaSP': sp.MaSP,
            'TenSP': sp.TenSP
        })
    return JsonResponse({'san_pham': data})

def them_chi_tiet_phieu_nhap(request, MaPhieuNhap):
    phieu_nhap = get_object_or_404(PhieuNhap, MaPhieuNhap=MaPhieuNhap)
    if phieu_nhap.TrangThai == 'completed':
        messages.error(request, "Phiếu nhập đã hoàn thành, không thể thêm chi tiết.")
        return redirect(reverse('UserMember:XemChiTietPhieuNhap', args=[MaPhieuNhap]))
    if request.method == 'POST':
        try:
            # Lấy dữ liệu từ form
            san_pham_id = request.POST.get('SanPham')
            so_luong_nhap = int(request.POST.get('SoLuongNhap'))
            gia_nhap = float(request.POST.get('GiaNhap'))
            ngay_san_xuat = datetime.strptime(request.POST.get('NgaySanXuat'), '%Y-%m-%d')
            ngay_het_han = datetime.strptime(request.POST.get('NgayHetHan'), '%Y-%m-%d')

            # Kiểm tra sản phẩm
            san_pham = get_object_or_404(SanPham, pk=san_pham_id)

            # **Ràng buộc NgaySanXuat và NgayHetHan**
            if ngay_het_han < ngay_san_xuat:
                messages.error(request, "Ngày hết hạn không được nhỏ hơn ngày sản xuất.")
                return redirect(reverse('UserMember:XemChiTietPhieuNhap', args=[MaPhieuNhap]))
            
            # Tạo chi tiết phiếu nhập
            chi_tiet = CT_PhieuNhap(
                PhieuNhap=phieu_nhap,
                SanPham=san_pham,
                SoLuongNhap=so_luong_nhap,
                GiaNhap=gia_nhap,
                NgaySanXuat=ngay_san_xuat,
                NgayHetHan=ngay_het_han
            )
            chi_tiet.save()
            
            # Cập nhật tổng số lượng và tổng tiền
            phieu_nhap.cap_nhat_tong_so_luong_va_tong_tien()
            messages.success(request, "Thêm chi tiết phiếu nhập thành công.")
            return redirect(reverse('UserMember:XemChiTietPhieuNhap', args=[MaPhieuNhap]))
        
        except ValueError:
            messages.error(request, "Dữ liệu không hợp lệ. Vui lòng kiểm tra lại.")
            return redirect(reverse('UserMember:XemChiTietPhieuNhap', args=[MaPhieuNhap]))
    
    return redirect(reverse('UserMember:XemChiTietPhieuNhap', args=[MaPhieuNhap]))
    
def sua_chi_tiet_phieu_nhap(request, MaPhieuNhap, MaCTPN):
    phieu_nhap = get_object_or_404(PhieuNhap, MaPhieuNhap=MaPhieuNhap)
    if phieu_nhap.TrangThai == 'completed':
        messages.error(request, "Phiếu nhập đã hoàn thành, không thể xóa chi tiết.")
        return redirect(reverse('UserMember:XemChiTietPhieuNhap', args=[MaPhieuNhap]))
    chi_tiet = get_object_or_404(CT_PhieuNhap, MaCTPN=MaCTPN)

    if request.method == 'POST':
        form = CT_PhieuNhapForm(request.POST, instance=chi_tiet)
        if form.is_valid():
            form.save()
        return redirect('UserMember:XemChiTietPhieuNhap', MaPhieuNhap=MaPhieuNhap)

    else:
        form =  CT_PhieuNhapForm(instance=chi_tiet)

    return render(request, 'sua_chi_tiet_phieu_nhap.html', {
        'form': form,
        'phieu_nhap': phieu_nhap,
        'chi_tiet': chi_tiet,
    })

def xoa_chi_tiet_phieu_nhap(request, MaPhieuNhap, MaCTPN):
    phieu_nhap = get_object_or_404(PhieuNhap, MaPhieuNhap=MaPhieuNhap)
    if phieu_nhap.TrangThai == 'completed':
        messages.error(request, "Phiếu nhập đã hoàn thành, không thể sửa chi tiết.")
        return redirect(reverse('UserMember:XemChiTietPhieuNhap', args=[MaPhieuNhap]))
    chi_tiet = get_object_or_404(CT_PhieuNhap, pk=MaCTPN, PhieuNhap__MaPhieuNhap=MaPhieuNhap)
    chi_tiet.delete()
    chi_tiet.PhieuNhap.cap_nhat_tong_so_luong_va_tong_tien()
    return redirect(reverse('UserMember:XemChiTietPhieuNhap', args=[MaPhieuNhap]))

def in_phieu_nhap(request, MaPhieuNhap):
    phieu_nhap = get_object_or_404(PhieuNhap, MaPhieuNhap=MaPhieuNhap)
    chi_tiet_phieu_nhap = CT_PhieuNhap.objects.filter(PhieuNhap=phieu_nhap)
    tong_tien_nhap = sum([item.SoLuongNhap * item.GiaNhap for item in chi_tiet_phieu_nhap])
    return render(request, 'pages/Admin/in_phieu_nhap.html', {
        'phieu_nhap': phieu_nhap,
        'chi_tiet_phieu_nhap': chi_tiet_phieu_nhap,
        'tong_tien_nhap': tong_tien_nhap,
    })
    
#---------------QL kho
@login_required(login_url='/DangNhap/')
@user_passes_test(is_admin, login_url='/DangNhap/')                         
def QLKho(request, simple=False):
    dskho_info = []
    if simple:
        # Lấy danh sách kho đơn giản, không tính toán thêm
        dskho_info = Kho.objects.all()
    else:
        # Lấy tất cả kho với tính toán chi tiết
        dskho = Kho.objects.all()
        for kho in dskho:
            so_loai_san_pham = TonKho.objects.filter(Kho=kho).values('SanPham').distinct().count()
            tong_so_luong_ton = TonKho.objects.filter(Kho=kho).aggregate(total=Sum(F('SoLuongTon')))['total'] or 0
            dskho_info.append({
                'kho': kho,
                'so_loai_san_pham': so_loai_san_pham,
                'tong_so_luong_ton': tong_so_luong_ton
            })
    
    return render(request, 'pages/Admin/QLKho.html', {'dskho_info': dskho_info})

def them_kho(request):
    if request.method == 'POST':
        form = KhoForm(request.POST)    
        if form.is_valid():
            form.save()
            messages.success(request, 'Kho mới đã được thêm thành công!')
            return redirect('UserMember:QLKho')  
        else:
            messages.error(request, 'Có lỗi xảy ra, vui lòng thử lại.')
    else:
        form = KhoForm()
    
    return render(request, 'pages/Admin/them_kho.html', {'form': form})     
def sua_kho(request):
    if request.method == 'POST':
        # Lấy mã kho từ form
        ma_kho = request.POST.get('MaKho')
        kho = get_object_or_404(Kho, MaKho=ma_kho)

        # Khởi tạo form với dữ liệu từ POST
        form = KhoForm(request.POST, instance=kho)
        
        # Kiểm tra tính hợp lệ của form
        if form.is_valid():
            form.save()
            messages.success(request, 'Kho đã được cập nhật thành công!')
            return redirect('UserMember:QLKho')
        else:
            messages.error(request, 'Cập nhật kho thất bại. Vui lòng kiểm tra lại dữ liệu.')
    
    else:
        ma_kho = request.GET.get('MaKho')  # Lấy mã kho từ URL
        kho = get_object_or_404(Kho, MaKho=ma_kho)
        form = KhoForm(instance=kho)

    return render(request, 'admin/edit_kho.html', {'form': form, 'kho': kho})     
def xoa_kho(request, kho_id):
    kho = get_object_or_404(Kho, pk=kho_id)
    if TonKho.objects.filter(Kho=kho).exists():
        messages.error(request, 'Không thể xóa kho này vì còn sản phẩm tồn.')
    else:
        kho.delete()
        messages.success(request, 'Kho đã được xóa thành công!')
    
    return redirect('UserMember:QLKho')  

from django.db.models import Q

def chi_tiet_ton_kho(request, kho_id):
    kho = get_object_or_404(Kho, pk=kho_id)
    keyword = request.GET.get("q", "").strip()
    # Lọc danh sách tồn kho dựa trên kho và từ khóa
    if keyword:
        # Giả sử bạn muốn tìm kiếm theo tên sản phẩm hoặc mã sản phẩm
        ton_kho_list = TonKho.objects.filter(Kho=kho).filter(
            Q(SanPham__TenSP__icontains=keyword) |  # Tìm theo tên sản phẩm
            Q(SanPham__MaSP__icontains=keyword)    # Tìm theo mã sản phẩm
        )
    else:
        # Nếu không có từ khóa, hiển thị toàn bộ danh sách tồn kho
        ton_kho_list = TonKho.objects.filter(Kho=kho)

    return render(request, 'pages/Admin/TonKho.html', {
        'kho': kho,
        'ton_kho_list': ton_kho_list,
        'keyword': keyword
    })
def them_ton_kho(phieu_nhap, kho):
    for ct in phieu_nhap.ct_phieunhap.all():
        TonKho.objects.create(
            SanPham=ct.SanPham,
            Kho=kho,
            SoLuongTon=ct.SoLuongNhap,
            NgayHetHan=ct.NgayHetHan,
            NgayNhapKho=phieu_nhap.NgayNhap
        )
def update_status(request):
    if request.method == 'POST':
        try:
            # Parse the JSON data from the request
            data = json.loads(request.body)
            MaPhieuNhap = data.get('MaPhieuNhap')
            new_status = data.get('TrangThai')
            kho_id = data.get('Kho')

            # Validate required fields
            if not MaPhieuNhap or not new_status:
                return JsonResponse({'success': False, 'message': 'Dữ liệu không đầy đủ.'})

            # Fetch the `PhieuNhap` object
            try:
                phieu_nhap = PhieuNhap.objects.get(MaPhieuNhap=MaPhieuNhap)
            except PhieuNhap.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'Phiếu nhập không tồn tại.'})

            # Check if the current status is already completed
            if phieu_nhap.TrangThai == 'completed':
                return JsonResponse({
                    'success': False,
                    'message': 'Không thể thay đổi trạng thái của phiếu nhập đã hoàn thành.'
                })

            # Handle the 'completed' status
            if new_status == 'completed':
                if not kho_id:
                    return JsonResponse({
                        'success': False,
                        'message': 'Vui lòng chọn kho để nhập hàng.'
                    })
                try:
                    kho = Kho.objects.get(MaKho=kho_id)
                except Kho.DoesNotExist:
                    return JsonResponse({'success': False, 'message': 'Kho không tồn tại.'})

                # Update stock when the status is completed
                them_ton_kho(phieu_nhap, kho)

            # Update and save the status
            phieu_nhap.TrangThai = new_status
            phieu_nhap.save()

            # Return success response with a flag to hide dropdown
            return JsonResponse({
                'success': True,
                'message': 'Trạng thái đã được cập nhật thành công!',
                'hide_status': new_status == 'completed'  # Flag để ẩn dropdown
            })

        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Dữ liệu JSON không hợp lệ.'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Đã xảy ra lỗi: {str(e)}'})

    return JsonResponse({'success': False, 'message': 'Yêu cầu không hợp lệ.'})

import openpyxl
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
from .models import TonKho
def xuat_excel_tonkho(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Danh sách sản phẩm trong kho"

    # Bước 2: Tiêu đề cột
    headers = [
        "STT", "Tên sản phẩm", "Tên kho", "Số lượng tồn", "Ngày nhập kho", "Ngày hết hạn", "Trạng thái"
    ]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Bước 3: Lấy dữ liệu từ cơ sở dữ liệu
    ton_kho_list = TonKho.objects.filter(IsDelete=False).select_related('SanPham', 'Kho')

    # Bước 4: Ghi dữ liệu vào sheet
    for idx, ton_kho in enumerate(ton_kho_list, start=1):
        sheet.cell(row=idx + 1, column=1, value=idx)  # STT
        sheet.cell(row=idx + 1, column=2, value=ton_kho.SanPham.TenSP)  # Tên sản phẩm
        sheet.cell(row=idx + 1, column=3, value=ton_kho.Kho.TenKho)  # Tên kho
        sheet.cell(row=idx + 1, column=4, value=ton_kho.SoLuongTon)  # Số lượng tồn
        sheet.cell(row=idx + 1, column=5, value=ton_kho.NgayNhapKho.strftime('%d/%m/%Y') if ton_kho.NgayNhapKho else "N/A")  # Ngày nhập kho
        sheet.cell(row=idx + 1, column=6, value=ton_kho.NgayHetHan.strftime('%d/%m/%Y') if ton_kho.NgayHetHan else "N/A")  # Ngày hết hạn
        sheet.cell(row=idx + 1, column=7, value="Còn hàng" if ton_kho.SoLuongTon > 0 else "Hết hàng")  # Trạng thái

    # Bước 5: Định dạng cột
    for col in sheet.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Lấy ký tự cột
        for cell in col:
            try:  # Tính độ dài tối đa của dữ liệu trong cột
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        sheet.column_dimensions[col_letter].width = adjusted_width

    # Bước 6: Trả file Excel về người dùng
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="Danh_sach_san_pham_trong_kho.xlsx"'
    workbook.save(response)

    return response

#-------------------QLKhachHang
@login_required(login_url='/DangNhap/')
@user_passes_test(is_admin, login_url='/DangNhap/')
def QLKhachHang(request):
    dsKhachHang = User.objects.filter(is_user=True).order_by('id')  # Lấy danh sách khách hàng
    for index, user in enumerate(dsKhachHang, start=1):
        user.stt_khachhang = index  # Đánh số thứ tự khách hàng
    return render(request, 'pages/Admin/QLKhachHang.html', {'dsKhachHang': dsKhachHang})
def get_user_details(request, user_id):
    user = get_object_or_404(User, pk=user_id)  # Tìm khách hàng theo id
    user_data = {
        'id': user.id,
        'username': user.username,
        'email': user.email,
        'phone': user.phone,
        'address': user.address,
        'full_name': user.full_name,
        'is_active': user.is_active,
    }
    return JsonResponse(user_data)

def sua_khachhang(request, user_id):
    user = get_object_or_404(User, pk=user_id)  # Lấy thông tin khách hàng từ ID
    if request.method == 'POST':
        form = KhachHangForm(request.POST, instance=user)  # Sử dụng KhachHangForm
        if form.is_valid():
            form.save()
            return redirect('UserMember:QLKhachHang')  # Quay lại danh sách khách hàng
    else:
        form = KhachHangForm(instance=user)  # Truyền dữ liệu khách hàng vào form
    return render(request, 'pages/Admin/suakhachhang.html', {'form': form, 'title': 'Sửa Khách Hàng'})


def update_stt_khachhang():
    user_list = User.objects.filter(is_user=True).order_by('id')  # Lấy danh sách khách hàng
    for index, user in enumerate(user_list, start=1):
        user.stt_khachhang = index
        user.save(update_fields=['stt_khachhang'])  

def danhsachdonhang_khachhang(request, user_id):
    khach_hang = get_object_or_404(CustomUser, pk=user_id, is_user=True)
    orders = DonHang.objects.filter(user=khach_hang).order_by('-NgayDat')  # Lấy danh sách đơn hàng
    return render(request, 'pages/Admin/DSDonHang_KhachHang.html', {
        'khach_hang': khach_hang,
        'orders': orders,  # Truyền orders sang template
    })
def xem_chitiet_donhang(request, order_id):
    don_hang = get_object_or_404(DonHang, pk=order_id)  # Lấy đơn hàng theo ID
    return render(request, 'pages/Admin/chitiet_don_hang.html', {
        'don_hang': don_hang,  # Truyền thông tin đơn hàng sang template
    })
#----------------------------Tin tức
@login_required(login_url='/DangNhap/')
@user_passes_test(is_admin, login_url='/DangNhap/')
def QLTinTuc(request): 
    dsTinTuc = TinTucModel.objects.all()  # Sử dụng tên mới TinTucModel
    return render(request, 'pages/Admin/QLTinTuc.html', {'dsTinTuc': dsTinTuc})

ALLOWED_IMAGE_EXTENSIONS = ['jpg', 'jpeg', 'png', 'gif', 'webp']
@csrf_exempt
def upload_image(request):
    if request.method == 'POST' and request.FILES.get('upload'):
        image = request.FILES['upload']
        path = default_storage.save('uploads/' + image.name, ContentFile(image.read()))
        url = default_storage.url(path)
        # Trả về phản hồi đúng định dạng CKEditor yêu cầu
        return JsonResponse({
            'uploaded': True,
            'url': url
        })
    return JsonResponse({'uploaded': False, 'error': {'message': 'No file uploaded'}}, status=400)
def them_tin_tuc(request):
    form = None
    if request.method == 'POST':
        form = TinTucForm(request.POST, request.FILES)
        if form.is_valid():
            tin_tuc = form.save(commit=False)  # Lưu tạm thời mà không commit vào DB ngay
            # Gán tác giả là admin nếu chưa có
            admin_user = get_user_model().objects.filter(is_admin=True).first()  # Tìm admin
            if admin_user:
                tin_tuc.TacGia = admin_user
                
            tin_tuc.save() 
            messages.success(request, 'Thêm tin tức thành công!')
            return redirect('UserMember:QLTinTuc')
        else:
            messages.error(request, 'Có lỗi trong quá trình thêm tin tức. Vui lòng kiểm tra lại.')
    if not form:
        form = TinTucForm()
    return render(request, 'pages/Admin/ThemTinTuc.html', {'form': form, 'title': 'Thêm Tin Tức'})

def save(self, *args, **kwargs):
        if self.order is None:
            last_order = TinTuc.objects.order_by('-order').first()
            self.order = last_order.order + 1 if last_order else 1
        super().save(*args, **kwargs)

def sua_tin_tuc(request, MaTin):
    # Sử dụng model TinTuc
    tin_tuc = get_object_or_404(TinTucModel, pk=MaTin)  # Đảm bảo rằng TinTuc là model đã import
    if request.method == 'POST':
        form = TinTucForm(request.POST, request.FILES, instance=tin_tuc)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cập nhật tin tức thành công!')
            return redirect('UserMember:QLTinTuc')
        else:
            messages.error(request, 'Có lỗi trong quá trình cập nhật. Vui lòng kiểm tra lại.')
    else:
        form = TinTucForm(instance=tin_tuc)
    
    return render(request, 'pages/Admin/SuaTinTuc.html', {'form': form, 'title': 'Sửa Tin Tức'})
def xoa_tin_tuc(request, MaTin):
    tin_tuc = get_object_or_404(TinTucModel, pk=MaTin)  # Đổi thành TinTucModel
    if request.method == 'POST':
        tin_tuc.delete()
        messages.success(request, 'Xóa tin tức thành công!')
        return redirect('UserMember:QLTinTuc')
    
    return render(request, 'pages/Admin/XoaTinTuc.html', {'tin_tuc': tin_tuc})
def xem_tin_tuc(request, MaTin):
    tin_tuc = get_object_or_404(TinTucModel, pk=MaTin)  # Đổi thành TinTucModel
    return render(request, 'pages/Admin/XemTinTuc.html', {'tin_tuc': tin_tuc})
    
#-----------------------------Thống kế

from django.db.models.functions import TruncDay, TruncMonth, TruncWeek, TruncYear
from django.db.models import Sum, F
from django.http import JsonResponse
from datetime import datetime
import calendar
from django.shortcuts import render

@login_required(login_url='/DangNhap/')
@user_passes_test(is_admin_or_staff_or_manager, login_url='/DangNhap/')
def thong_ke(request):
    today = datetime.today().date()
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    doanh_thu_khoang_thoi_gian = 0
    errors = []

    # Doanh thu trong ngày hôm nay
    doanh_thu_ngay = DonHang.objects.filter(NgayDat=today).aggregate(
        tong_doanh_thu=Sum('TongTien')
    )['tong_doanh_thu'] or 0

    # Doanh thu trong khoảng thời gian
    if start_date and end_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()

            # Kiểm tra chỉ khi không thể đảm bảo dữ liệu từ client
            if start_date_obj > date.today() or end_date_obj > date.today() or start_date_obj > end_date_obj:
                errors.append("Dữ liệu ngày không hợp lệ.")
            else:
                # Tính doanh thu trong khoảng thời gian hợp lệ
                doanh_thu_khoang_thoi_gian = DonHang.objects.filter(
                    NgayDat__range=[start_date_obj, end_date_obj]
                ).aggregate(tong_doanh_thu=Sum('TongTien'))['tong_doanh_thu'] or 0
        except ValueError:
            errors.append("Định dạng ngày không hợp lệ.")

    # Thống kê voucher
    voucher_ap_dung = Voucher.objects.filter(
        SoLuong__gt=0, NgayKetThuc__gte=today, NgayBatDau__lte=today
    )
    tong_so_voucher = voucher_ap_dung.aggregate(
        tong_voucher=Sum('SoLuong')
    )['tong_voucher'] or 0
    tong_voucher_ap_dung = DonHang.objects.filter(
        voucher__isnull=False,  # Chỉ lấy các đơn hàng có áp dụng voucher
        NgayDat__lte=today,    # Ngày đặt nhỏ hơn hoặc bằng hôm nay
    ).aggregate(
        tong_giam=Sum(F('TongTien') * F('voucher__PhanTramGiam') / 100)
    )['tong_giam'] or 0

    tong_voucher_ap_dung = round(tong_voucher_ap_dung, 3)
    tong_so_luong_da_ban = CT_DonHang.objects.aggregate(
    tong_so_luong=Sum('SoLuong')
    )['tong_so_luong'] or 0 
    # Thống kê sản phẩm đã bán
    loai_san_pham_ban = CT_DonHang.objects.values(
        'MaSP__MaLoai__TenLoai'
    ).annotate(
        tong_ban=Sum('SoLuong')
    ).order_by('-tong_ban')

    # Top 3 sản phẩm bán chạy
    top_san_pham = CT_DonHang.objects.values(
        'MaSP__TenSP'
    ).annotate(
        tong_ban=Sum('SoLuong')
    ).order_by('-tong_ban')[:3]

    # Thống kê tồn kho
    tong_so_luong_ton_kho = TonKho.objects.aggregate(
        tong_so_luong=Sum('SoLuongTon')
    )['tong_so_luong'] or 0
    tong_gia_tri_ton_kho = TonKho.objects.annotate(
        gia_tri=F('SoLuongTon') * F('SanPham__GiaBan')
    ).aggregate(
        tong_gia_tri=Sum('gia_tri')
    )['tong_gia_tri'] or 0

    # Sản phẩm tồn kho lâu ngày
    san_pham_ton_lau = TonKho.objects.filter(
        NgayNhapKho__lt=today - timedelta(days=15)
    ).select_related('SanPham', 'Kho')
    inventory_summary = get_inventory_summary()
    # Trả về template
    return render(request, 'pages/admin/ThongKe.html', {
        'doanh_thu_ngay': doanh_thu_ngay,
        'doanh_thu_khoang_thoi_gian': doanh_thu_khoang_thoi_gian,
        'start_date': start_date,
        'end_date': end_date,
        'tong_so_voucher': tong_so_voucher,
        'tong_voucher_ap_dung': tong_voucher_ap_dung,
        'loai_san_pham_ban': loai_san_pham_ban,
        'tong_so_luong_da_ban': tong_so_luong_da_ban,
        'tong_so_luong_ton_kho': tong_so_luong_ton_kho,
        'tong_gia_tri_ton_kho': tong_gia_tri_ton_kho,
        'san_pham_ton_lau': san_pham_ton_lau,
        'top_san_pham': top_san_pham,  # Thêm vào context
        'errors': errors,
        'warehouses': inventory_summary["warehouses"],
        'table_data': inventory_summary["table_data"],
    })
def get_inventory_summary():
    # Lấy danh sách kho và sản phẩm
    warehouses = Kho.objects.all()
    products = SanPham.objects.all().order_by('TenSP')

    # Chuẩn bị dữ liệu bảng
    table_data = []
    for product in products:
        row = {
            "TenSP": product.TenSP,
            "Warehouses": [],
            "TongSoLuong": 0,
        }
        for warehouse in warehouses:
            quantity = TonKho.objects.filter(
                SanPham=product, Kho=warehouse, IsDelete=False
            ).aggregate(total=Sum('SoLuongTon'))['total'] or 0
            row["Warehouses"].append({"Kho": warehouse.TenKho, "SoLuong": quantity})
            row["TongSoLuong"] += quantity
        table_data.append(row)

    return {
        "warehouses": [kho.TenKho for kho in Kho.objects.all()],
        "table_data": table_data,
    }

#--------Bieu Do Doanh Thu
def thong_ke_bieu_do(request):
    # Determine the type of chart ('day', 'week', 'month', 'year') and current_date
    chart_type = request.GET.get('type', 'year')
    current_date_str = request.GET.get('current_date', None)
    if current_date_str:
        try:
            current_date = datetime.strptime(current_date_str, '%Y-%m-%d').date()
        except ValueError:
            current_date = datetime.today().date()  # Giá trị mặc định nếu ngày không hợp lệ
    else:
        current_date = datetime.today().date()
    current_date = datetime.strptime(current_date_str, '%Y-%m-%d').date()

    labels, values = [], []

    if chart_type == 'day':
        last_7_days = [current_date - timedelta(days=i) for i in range(6, -1, -1)]
        labels = [day.strftime('%Y-%m-%d') for day in last_7_days]
        data = DonHang.objects.annotate(period=TruncDay('NgayDat')).values('period').annotate(tong_doanh_thu=Sum('TongTien'))
        data_dict = {
        (item['period'].strftime('%Y-%m-%d') if item['period'] else 'N/A'): item['tong_doanh_thu'] or 0
    for item in data
        }
        values = [data_dict.get(day, 0) for day in labels]

    elif chart_type == 'week':
        start_of_week = current_date - timedelta(days=current_date.weekday())
        last_8_weeks = [start_of_week - timedelta(weeks=i) for i in range(7, -1, -1)]
        labels = [
        f"{(week - timedelta(days=6)).strftime('%d/%m/%Y')} - {week.strftime('%d/%m/%Y')}"
        for week in last_8_weeks]
        data = DonHang.objects.annotate(period=TruncWeek('NgayDat')).values('period').annotate(tong_doanh_thu=Sum('TongTien'))
        data_dict = {
        (item['period'].strftime('%Y-%U') if item['period'] else 'N/A'): item['tong_doanh_thu'] or 0
        for item in data
    }
        values = [data_dict.get(week.strftime('%Y-%U'), 0) for week in last_8_weeks]

    elif chart_type == 'month':
        labels = [f"{i}/{current_date.year}" for i in range(1, 13)]  # Tháng/Năm
        data = DonHang.objects.annotate(period=TruncMonth('NgayDat')).values('period').annotate(tong_doanh_thu=Sum('TongTien'))
        data_dict = {
            (item['period'].strftime('%Y-%m') if item['period'] else 'N/A'): item['tong_doanh_thu'] or 0
            for item in data
        }
        values = [data_dict.get(f"{current_date.year}-{str(i).zfill(2)}", 0) for i in range(1, 13)]

    elif chart_type == 'year':
        labels = [str(current_date.year - i) for i in range(4, -1, -1)]
        data = DonHang.objects.annotate(period=TruncYear('NgayDat')).values('period').annotate(tong_doanh_thu=Sum('TongTien'))
        data_dict = {
            (item['period'].strftime('%Y') if item['period'] else 'N/A'): item['tong_doanh_thu'] or 0
            for item in data
        }
        values = [data_dict.get(year, 0) for year in labels]
    return JsonResponse({'labels': labels, 'values': values})

#--------Bieu Do Voucher
def thong_ke_voucher_bieu_do(request):
    chart_type = request.GET.get('type', 'year')
    current_date_str = request.GET.get('current_date', None)

    if current_date_str:
        try:
            current_date = datetime.strptime(current_date_str, '%Y-%m-%d').date()
        except ValueError:
            current_date = datetime.today().date()
    else:
        current_date = datetime.today().date()

    labels, values = [], []

    if chart_type == 'day':
        last_7_days = [current_date - timedelta(days=i) for i in range(6, -1, -1)]
        labels = [day.strftime('%Y-%m-%d') for day in last_7_days]
        data = DonHang.objects.annotate(period=TruncDay('NgayDat')).values('period').annotate(
            total_discount=Sum(F('TongTien') * F('voucher__PhanTramGiam') / 100)
        )
        data_dict = {
            (item['period'].strftime('%Y-%m-%d') if item['period'] else 'N/A'): item['total_discount'] or 0
            for item in data
        }
        values = [data_dict.get(day, 0) for day in labels]

    elif chart_type == 'week':
        start_of_week = current_date - timedelta(days=current_date.weekday())
        last_8_weeks = [start_of_week - timedelta(weeks=i) for i in range(7, -1, -1)]
        labels = [
        f"{(week - timedelta(days=6)).strftime('%d/%m/%Y')} - {week.strftime('%d/%m/%Y')}"
        for week in last_8_weeks]
        data = DonHang.objects.annotate(period=TruncWeek('NgayDat')).values('period').annotate(
            total_discount=Sum(F('TongTien') * F('voucher__PhanTramGiam') / 100)
        )
        data_dict = {
            (item['period'].strftime('%Y-%U') if item['period'] else 'N/A'): item['total_discount'] or 0
            for item in data
        }
        values = [data_dict.get(week.strftime('%Y-%U'), 0) for week in last_8_weeks]

    elif chart_type == 'month':
        labels = [f"{i}/{current_date.year}" for i in range(1, 13)]
        data = DonHang.objects.annotate(period=TruncMonth('NgayDat')).values('period').annotate(
            total_discount=Sum(F('TongTien') * F('voucher__PhanTramGiam') / 100)
        )
        data_dict = {
            (item['period'].strftime('%Y-%m') if item['period'] else 'N/A'): item['total_discount'] or 0
            for item in data
        }
        values = [data_dict.get(f"{current_date.year}-{str(i).zfill(2)}", 0) for i in range(1, 13)]

    elif chart_type == 'year':
        labels = [str(current_date.year - i) for i in range(4, -1, -1)]
        data = DonHang.objects.annotate(period=TruncYear('NgayDat')).values('period').annotate(
            total_discount=Sum(F('TongTien') * F('voucher__PhanTramGiam') / 100)
        )
        data_dict = {
            (item['period'].strftime('%Y') if item['period'] else 'N/A'): item['total_discount'] or 0
            for item in data
        }
        values = [data_dict.get(year, 0) for year in labels]

    return JsonResponse({'labels': labels, 'values': values})

#--------Bieu Do San Pham
def thong_ke_san_pham_bieu_do(request):
    chart_type = request.GET.get('type', 'year')
    current_date_str = request.GET.get('current_date', None)

    if current_date_str:
        try:
            current_date = datetime.strptime(current_date_str, '%Y-%m-%d').date()
        except ValueError:
            current_date = datetime.today().date()
    else:
        current_date = datetime.today().date()

    labels, values = [], []

    if chart_type == 'day':
        # 7 ngày gần đây
        last_7_days = [current_date - timedelta(days=i) for i in range(6, -1, -1)]
        labels = [day.strftime('%Y-%m-%d') for day in last_7_days]
        data = CT_DonHang.objects.filter(gio_hang__NgayDat__isnull=False).annotate(
            period=TruncDay('gio_hang__NgayDat')
        ).values('period').annotate(tong_san_pham=Sum('SoLuong'))
        data_dict = {
            (item['period'].strftime('%Y-%m-%d') if item['period'] else 'N/A'): item['tong_san_pham'] or 0
            for item in data
        }
        values = [data_dict.get(day, 0) for day in labels]

    elif chart_type == 'week':
        # 8 tuần gần đây
        start_of_week = current_date - timedelta(days=current_date.weekday())
        last_8_weeks = [start_of_week - timedelta(weeks=i) for i in range(7, -1, -1)]
        labels = [
        f"{(week - timedelta(days=6)).strftime('%d/%m/%Y')} - {week.strftime('%d/%m/%Y')}"
        for week in last_8_weeks]
        data = CT_DonHang.objects.filter(gio_hang__NgayDat__isnull=False).annotate(
            period=TruncWeek('gio_hang__NgayDat')
        ).values('period').annotate(tong_san_pham=Sum('SoLuong'))
        data_dict = {
            (item['period'].strftime('%Y-%U') if item['period'] else 'N/A'): item['tong_san_pham'] or 0
            for item in data
        }
        values = [data_dict.get(week.strftime('%Y-%U'), 0) for week in last_8_weeks]

    elif chart_type == 'month':
        # 12 tháng trong năm hiện tại
        labels = [f"{i}/{current_date.year}" for i in range(1, 13)]
        data = CT_DonHang.objects.filter(gio_hang__NgayDat__isnull=False).annotate(
            period=TruncMonth('gio_hang__NgayDat')
        ).values('period').annotate(tong_san_pham=Sum('SoLuong'))
        data_dict = {
            (item['period'].strftime('%Y-%m') if item['period'] else 'N/A'): item['tong_san_pham'] or 0
            for item in data
        }
        values = [data_dict.get(f"{current_date.year}-{str(i).zfill(2)}", 0) for i in range(1, 13)]

    elif chart_type == 'year':
        # 5 năm gần đây
        labels = [str(current_date.year - i) for i in range(4, -1, -1)]
        data = CT_DonHang.objects.filter(gio_hang__NgayDat__isnull=False).annotate(
            period=TruncYear('gio_hang__NgayDat')
        ).values('period').annotate(tong_san_pham=Sum('SoLuong'))
        data_dict = {
            (item['period'].strftime('%Y') if item['period'] else 'N/A'): item['tong_san_pham'] or 0
            for item in data
        }
        values = [data_dict.get(year, 0) for year in labels]

    return JsonResponse({'labels': labels, 'values': values})

#-------------------danh giá
def DanhGias(request, MaDH):
    don_hang = get_object_or_404(DonHang, MaDH=MaDH, user=request.user)
    if don_hang.TrangThai != 'da_hoan_thanh':
        return render(request, 'error.html', {'message': 'Chỉ có thể đánh giá đơn hàng đã hoàn thành.'})

    # Lấy tên người dùng
    ten_nguoi_dung = request.user.username

    if request.method == 'POST':
        form = DanhGiaForm(request.POST, request.FILES)
        if form.is_valid():
            danh_gia = form.save(commit=False)
            danh_gia.MaDH = don_hang
            danh_gia.TKViet = don_hang.user.username
            danh_gia.Rating = form.cleaned_data['Rating']  # Gán Mức Độ Hài Lòng
            danh_gia.save()  # Lưu đối tượng vào cơ sở dữ liệu
            return redirect('UserMember:LSMuaHang')
        else:
            print("Form errors:", form.errors)
    else:
        form = DanhGiaForm()

    return render(request, 'pages/User/DanhGia.html', {
        'form': form,
        'don_hang': don_hang,
        'ten_nguoi_dung': ten_nguoi_dung,
    })

@login_required(login_url='/DangNhap/')
@user_passes_test(is_admin, login_url='/DangNhap/')
def QLDanhGia(request):
    danh_gias = DanhGia.objects.all().order_by('NgayViet')
    return render(request, 'pages/Admin/QLDanhGia.html', {'danh_gias': danh_gias})